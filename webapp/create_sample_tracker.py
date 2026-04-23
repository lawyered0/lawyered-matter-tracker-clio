#!/usr/bin/env python3
"""Create a sample matter-tracker.xlsx with the correct schema and demo data."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
import os

HEADERS = [
    "File #", "Client Name", "Matter Description", "Status", "Date Opened",
    "Date Closed", "Last Activity", "Opposing Party", "Next Action / Deadline",
    "Timeline", "Client ID Verified", "Conflict Check Done", "Client Email",
    "Client Phone", "Client Address", "Discovery Date", "Limitation Statute",
    "Limitation Deadline", "Court Deadlines", "Matter Folder",
    "Other Parties / Related Persons", "Matter Type",
]

COL_WIDTHS = {
    1: 12, 2: 22, 3: 40, 4: 10, 5: 14, 6: 14, 7: 14, 8: 22, 9: 30,
    10: 60, 11: 14, 12: 14, 13: 22, 14: 16, 15: 30, 16: 14, 17: 30,
    18: 14, 19: 40, 20: 50, 21: 50, 22: 16,
}

WRAP_COLS = {3, 9, 10, 19, 21}  # C, I, J, S, U

HEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=10, bold=True)
DATA_FONT = Font(name="Arial", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

SAMPLE_MATTERS = [
    {
        "File #": "2026-001",
        "Client Name": "Smith, John",
        "Matter Description": "Employment termination — wrongful dismissal claim",
        "Status": "Open",
        "Date Opened": "2026-01-15",
        "Last Activity": "2026-03-10",
        "Opposing Party": "Acme Corp",
        "Next Action / Deadline": "2026-04-01: File statement of claim",
        "Timeline": "SUMMARY: John Smith was terminated without cause from Acme Corp after 8 years. Severance offer of 4 weeks; seeking 8 months. Demand letter sent; no response.\n\n2026-01-15: Client retained re: wrongful dismissal\n2026-01-20: Reviewed employment contract and termination letter\n2026-02-01: Sent demand letter to Acme Corp\n2026-03-10: Follow-up sent; no response from employer",
        "Client ID Verified": "✓",
        "Conflict Check Done": "✓",
        "Client Email": "john.smith@example.com",
        "Client Phone": "(416) 555-0101",
        "Discovery Date": "2026-01-10",
        "Limitation Statute": "general_statute",
        "Limitation Deadline": "2028-01-10",
        "Matter Type": "Employment",
    },
    {
        "File #": "2026-002",
        "Client Name": "Rivera Holdings Inc. (Maria Rivera)",
        "Matter Description": "Commercial lease dispute — landlord withholding deposit",
        "Status": "Open",
        "Date Opened": "2026-02-01",
        "Last Activity": "2026-03-05",
        "Opposing Party": "Downtown Properties Ltd.",
        "Next Action / Deadline": "2026-03-25: Settlement conference at 10:00 AM",
        "Timeline": "SUMMARY: Rivera Holdings paid $15,000 deposit for commercial lease at 100 Main St. Landlord terminated early and refuses to return deposit. Small Claims Court claim filed.\n\n2026-02-01: Client retained re: deposit recovery\n2026-02-10: Filed Plaintiff's Claim in Small Claims Court\n2026-03-05: Defence received; settlement conference scheduled",
        "Client ID Verified": "✓",
        "Conflict Check Done": "✓",
        "Client Email": "maria@riveraholdings.example.com",
        "Client Phone": "(416) 555-0202",
        "Discovery Date": "2026-01-15",
        "Limitation Statute": "general_statute",
        "Limitation Deadline": "2028-01-15",
        "Court Deadlines": '[{"date":"2026-03-25","description":"Settlement conference","source":"Notice of Settlement Conference"}]',
        "Matter Type": "SCC",
    },
    {
        "File #": "2026-003",
        "Client Name": "Chen, David",
        "Matter Description": "Share purchase agreement — acquisition of tech startup",
        "Status": "Open",
        "Date Opened": "2026-02-20",
        "Last Activity": "2026-03-15",
        "Opposing Party": "",
        "Next Action / Deadline": "Review final SPA draft and send to client",
        "Timeline": "SUMMARY: David Chen acquiring 60% stake in TechStartup Inc. for $500K. Due diligence complete. SPA in final draft stage.\n\n2026-02-20: Client retained re: share purchase\n2026-02-28: Completed due diligence review\n2026-03-10: First draft SPA circulated\n2026-03-15: Received comments from seller's counsel",
        "Client ID Verified": "✓",
        "Conflict Check Done": "✓",
        "Client Email": "david.chen@example.com",
        "Matter Type": "Corporate",
    },
]


def create_sheet(wb, name, data_rows=None, tab_color=None):
    ws = wb.create_sheet(title=name)
    if tab_color:
        ws.sheet_properties.tabColor = tab_color

    # Header row
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(wrap_text=(col in WRAP_COLS))

    # Column widths
    for col, width in COL_WIDTHS.items():
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    # Freeze panes and auto-filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(HEADERS)).column_letter}1"

    # Data validation
    status_dv = DataValidation(type="list", formula1='"Open,Closed"', allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add(f"D2:D1000")

    statute_dv = DataValidation(
        type="list",
        formula1='"general_statute,general_ultimate,consumer_protection,employment,civil_rights,construction,custom"',
        allow_blank=True,
    )
    ws.add_data_validation(statute_dv)
    statute_dv.add(f"Q2:Q1000")

    # Data rows
    if data_rows:
        for row_idx, matter in enumerate(data_rows, 2):
            for col, header in enumerate(HEADERS, 1):
                val = matter.get(header, "")
                cell = ws.cell(row=row_idx, column=col, value=val if val else None)
                cell.font = DATA_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(wrap_text=(col in WRAP_COLS))

    return ws


def main():
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    create_sheet(wb, "Open Matters", data_rows=SAMPLE_MATTERS)
    create_sheet(wb, "Closed Matters", tab_color="808080")

    out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "matter-tracker.xlsx")
    wb.save(out_path)
    print(f"Created sample tracker: {out_path}")
    print(f"  {len(SAMPLE_MATTERS)} sample matters added to 'Open Matters'")
    print(f"  'Closed Matters' sheet ready (empty)")


if __name__ == "__main__":
    main()
