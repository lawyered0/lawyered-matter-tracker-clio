#!/usr/bin/env python3
"""Solo Law Matter Tracker — local web app backed by Excel."""

import os
import re
import json
import shutil
import subprocess
from datetime import datetime, date, timedelta
from flask import Flask, render_template, jsonify, request, send_from_directory
from openpyxl import load_workbook

app = Flask(__name__, static_folder="static")
XLSX_PATH = os.path.join(os.path.dirname(__file__), "matter-tracker.xlsx")
BACKUP_DIR = os.path.join(os.path.dirname(__file__), "backups")
# Set this to the root folder where your client matter folders live
MATTER_FOLDER_BASE = os.environ.get(
    "MATTER_FOLDER_BASE",
    os.path.join(os.path.dirname(__file__), "matters"),
)


def backup_spreadsheet():
    """Create a timestamped backup before any write operation."""
    os.makedirs(BACKUP_DIR, exist_ok=True)
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    backup_path = os.path.join(BACKUP_DIR, f"matter-tracker_{timestamp}.xlsx")
    shutil.copy2(XLSX_PATH, backup_path)
    # Keep only the 20 most recent backups
    backups = sorted(
        [f for f in os.listdir(BACKUP_DIR) if f.endswith(".xlsx")],
        reverse=True,
    )
    for old in backups[20:]:
        os.remove(os.path.join(BACKUP_DIR, old))

MATTER_TYPES = ["SCC", "Employment", "Demand/Recovery", "Corporate", "Transactional", "Other"]


def ensure_column_exists(column_name):
    """Add a column header to both sheets if it doesn't already exist."""
    wb = load_workbook(XLSX_PATH)
    modified = False
    for sheet_name in ["Open Matters", "Closed Matters"]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        if column_name not in headers:
            new_col = len(headers) + 1
            ws.cell(row=1, column=new_col).value = column_name
            modified = True
    if modified:
        backup_spreadsheet()
        wb.save(XLSX_PATH)
    wb.close()


# Ensure "Matter Type" column exists on startup
try:
    ensure_column_exists("Matter Type")
except Exception:
    pass  # Don't block startup if spreadsheet is locked


# Limitation statutes — customise these for your jurisdiction
LIMITATION_STATUTES = {
    "general_statute": {"name": "General statute of limitations — 2-year", "years": 2},
    "general_ultimate": {"name": "Ultimate repose — 10-year", "years": 10},
    "consumer_protection": {"name": "Consumer protection — 2-year", "years": 2},
    "employment": {"name": "Employment claims — 2-year", "years": 2},
    "civil_rights": {"name": "Civil rights — 1-year", "years": 1},
    "construction": {"name": "Construction defect — 2-year", "years": 2},
    "custom": {"name": "Custom (manual deadline)", "years": None},
}


def parse_date(val):
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    return s if s else None


def parse_timeline(val):
    if not val:
        return []
    entries = []
    for line in str(val).split("\n"):
        line = line.strip()
        if not line:
            continue
        if len(line) >= 10 and line[4] == "-" and line[7] == "-":
            entries.append({"date": line[:10], "text": line[11:].strip().lstrip(":").strip()})
        else:
            entries.append({"date": None, "text": line})
    return entries


def parse_json_field(val):
    if not val:
        return []
    s = str(val).strip()
    if s.startswith("["):
        try:
            return json.loads(s)
        except Exception:
            pass
    return []


def calc_limitation_deadline(discovery_date_str, statute_key):
    if not discovery_date_str or not statute_key:
        return None
    statute = LIMITATION_STATUTES.get(statute_key)
    if not statute or not statute["years"]:
        return None
    try:
        d = datetime.strptime(discovery_date_str, "%Y-%m-%d")
        target_year = d.year + statute["years"]
        # Handle leap year edge case (Feb 29 → Feb 28)
        try:
            return d.replace(year=target_year).strftime("%Y-%m-%d")
        except ValueError:
            return d.replace(year=target_year, day=28).strftime("%Y-%m-%d")
    except (ValueError, OverflowError):
        return None


def load_matters():
    try:
        wb = load_workbook(XLSX_PATH, data_only=True)
    except Exception as e:
        app.logger.error(f"Failed to open spreadsheet: {e}")
        return []
    matters = []
    for sheet_name in ["Open Matters", "Closed Matters"]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        try:
            rows = list(ws.iter_rows(min_row=1, values_only=True))
        except Exception as e:
            app.logger.error(f"Failed to read sheet '{sheet_name}': {e}")
            continue
        if not rows:
            continue
        headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
        for row in rows[1:]:
            if not row or not row[0]:
                continue
            record = {}
            for i, h in enumerate(headers):
                record[h] = row[i] if i < len(row) else None
            record["_sheet"] = sheet_name
            matters.append(record)
    wb.close()
    return matters


def safe_str(val, default=""):
    """Safely convert any cell value to string."""
    if val is None:
        return default
    try:
        return str(val)
    except Exception:
        return default


def matter_to_dict(m):
    try:
        return _matter_to_dict(m)
    except Exception as e:
        file_no = safe_str(m.get("File #"), "???")
        app.logger.error(f"Error parsing matter {file_no}: {e}")
        return {
            "file_no": file_no,
            "client": safe_str(m.get("Client Name")),
            "description": safe_str(m.get("Matter Description")),
            "status": safe_str(m.get("Status")),
            "date_opened": None, "date_closed": None, "last_activity": None,
            "opposing": "", "next_action": "", "deadline_date": None,
            "earliest_deadline": None, "timeline": [], "timeline_raw": "",
            "id_verified": "", "conflict_check": "", "matter_type": "Other",
            "sheet": m.get("_sheet", ""), "client_email": "", "client_phone": "",
            "client_address": "", "discovery_date": None, "limitation_statute": "",
            "limitation_deadline": None, "has_limitation": False,
            "court_deadlines": [], "all_deadlines": [],
            "matter_folder": "",
            "other_parties": "",
            "_parse_error": str(e),
        }


def _matter_to_dict(m):
    next_action = safe_str(m.get("Next Action / Deadline"))
    date_opened = parse_date(m.get("Date Opened"))
    date_closed = parse_date(m.get("Date Closed"))
    last_activity = parse_date(m.get("Last Activity"))
    discovery_date = parse_date(m.get("Discovery Date"))
    limitation_statute = safe_str(m.get("Limitation Statute"))
    limitation_deadline = parse_date(m.get("Limitation Deadline"))

    # Only calculate if explicitly set
    if discovery_date and limitation_statute and not limitation_deadline:
        limitation_deadline = calc_limitation_deadline(discovery_date, limitation_statute)

    # Has limitation tracking been enabled?
    has_limitation = bool(discovery_date or limitation_statute or limitation_deadline)

    manual_type = safe_str(m.get("Matter Type")).strip()
    if manual_type and manual_type in MATTER_TYPES:
        matter_type = manual_type
    else:
        desc = safe_str(m.get("Matter Description")).lower()
        if "small claims" in desc or "sc-" in desc:
            matter_type = "SCC"
        elif "employment" in desc or "roe" in desc or "wrongful" in desc or "termination" in desc:
            matter_type = "Employment"
        elif "demand" in desc or "debt recovery" in desc:
            matter_type = "Demand/Recovery"
        elif "shareholder" in desc or "share purchase" in desc or "corporate" in desc:
            matter_type = "Corporate"
        elif "contract" in desc or "lease" in desc:
            matter_type = "Transactional"
        else:
            matter_type = "Other"

    # Parse deadline from next_action text
    deadline_date = None
    if next_action:
        date_matches = re.findall(r'\b(\d{4}-\d{2}-\d{2})\b', next_action)
        for dm in date_matches:
            try:
                datetime.strptime(dm, "%Y-%m-%d")
                deadline_date = dm
                break
            except ValueError:
                pass

    # Court deadlines — simple manual entries
    court_deadlines = parse_json_field(m.get("Court Deadlines"))

    # Collect ALL dates that matter for dashboard urgency
    all_deadlines = []
    if deadline_date:
        all_deadlines.append({"date": deadline_date, "label": "Next action", "type": "action"})
    if limitation_deadline:
        all_deadlines.append({"date": limitation_deadline, "label": "Limitation expires", "type": "limitation"})
    for cd in court_deadlines:
        if cd.get("date"):
            all_deadlines.append({"date": cd["date"], "label": cd.get("description", "Court deadline"), "type": "court"})

    # Earliest upcoming deadline for sorting
    earliest_deadline = None
    for dl in all_deadlines:
        if not earliest_deadline or dl["date"] < earliest_deadline:
            earliest_deadline = dl["date"]

    return {
        "file_no": safe_str(m.get("File #")),
        "client": safe_str(m.get("Client Name")),
        "description": safe_str(m.get("Matter Description")),
        "status": safe_str(m.get("Status")),
        "date_opened": date_opened,
        "date_closed": date_closed,
        "last_activity": last_activity,
        "opposing": safe_str(m.get("Opposing Party")),
        "next_action": next_action,
        "deadline_date": deadline_date,
        "earliest_deadline": earliest_deadline,
        "timeline": parse_timeline(m.get("Timeline", "")),
        "timeline_raw": safe_str(m.get("Timeline")),
        "id_verified": safe_str(m.get("Client ID Verified")),
        "conflict_check": safe_str(m.get("Conflict Check Done")),
        "matter_type": matter_type,
        "sheet": m.get("_sheet", ""),
        "client_email": safe_str(m.get("Client Email")),
        "client_phone": safe_str(m.get("Client Phone")),
        "client_address": safe_str(m.get("Client Address")),
        "discovery_date": discovery_date,
        "limitation_statute": limitation_statute,
        "limitation_deadline": limitation_deadline,
        "has_limitation": has_limitation,
        "court_deadlines": court_deadlines,
        "all_deadlines": all_deadlines,
        "matter_folder": safe_str(m.get("Matter Folder")),
        "other_parties": safe_str(m.get("Other Parties / Related Persons")),
    }


def write_cell(file_no, column_name, value):
    return write_cells(file_no, {column_name: value})


def write_cells(file_no, updates):
    """Write multiple column values for a single matter in one save operation."""
    try:
        backup_spreadsheet()
        wb = load_workbook(XLSX_PATH)
    except Exception as e:
        app.logger.error(f"Failed to open spreadsheet for writing: {e}")
        return False
    try:
        for sheet_name in ["Open Matters", "Closed Matters"]:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            headers = [cell.value for cell in ws[1]]
            file_col = headers.index("File #") + 1 if "File #" in headers else None
            if not file_col:
                continue
            col_map = {}
            for col_name in updates:
                if col_name in headers:
                    col_map[col_name] = headers.index(col_name) + 1
            if not col_map:
                continue
            for row in range(2, ws.max_row + 1):
                if str(ws.cell(row=row, column=file_col).value) == str(file_no):
                    for col_name, col_idx in col_map.items():
                        ws.cell(row=row, column=col_idx).value = updates[col_name]
                    wb.save(XLSX_PATH)
                    wb.close()
                    return True
    except Exception as e:
        app.logger.error(f"Failed to write cells [{file_no}]: {e}")
    wb.close()
    return False


def add_timeline_entry(file_no, entry_date, entry_text):
    try:
        backup_spreadsheet()
        wb = load_workbook(XLSX_PATH)
    except Exception as e:
        app.logger.error(f"Failed to open spreadsheet for timeline write: {e}")
        return False
    try:
        for sheet_name in ["Open Matters", "Closed Matters"]:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            headers = [cell.value for cell in ws[1]]
            file_col = headers.index("File #") + 1 if "File #" in headers else None
            timeline_col = headers.index("Timeline") + 1 if "Timeline" in headers else None
            activity_col = headers.index("Last Activity") + 1 if "Last Activity" in headers else None
            if not file_col or not timeline_col:
                continue
            for row in range(2, ws.max_row + 1):
                if str(ws.cell(row=row, column=file_col).value) == str(file_no):
                    current = ws.cell(row=row, column=timeline_col).value or ""
                    ws.cell(row=row, column=timeline_col).value = str(current) + f"\n{entry_date}: {entry_text}"
                    if activity_col:
                        ws.cell(row=row, column=activity_col).value = entry_date
                    wb.save(XLSX_PATH)
                    wb.close()
                    return True
    except Exception as e:
        app.logger.error(f"Failed to add timeline entry for {file_no}: {e}")
    wb.close()
    return False


# --- Routes ---

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/logo.png")
def logo():
    logo_path = os.path.join(os.path.dirname(__file__), "logo.png")
    if os.path.exists(logo_path):
        return send_from_directory(os.path.dirname(__file__), "logo.png")
    # Return a 1x1 transparent pixel if no logo exists
    from flask import Response
    return Response(
        b'\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01\x08\x06\x00\x00\x00\x1f\x15\xc4\x89\x00\x00\x00\nIDATx\x9cc\x00\x01\x00\x00\x05\x00\x01\r\n\xb4\x00\x00\x00\x00IEND\xaeB`\x82',
        mimetype='image/png'
    )


@app.route("/api/matters")
def api_matters():
    try:
        return jsonify([matter_to_dict(m) for m in load_matters()])
    except Exception as e:
        app.logger.error(f"Failed to load matters: {e}")
        return jsonify({"error": "Failed to load matters", "detail": str(e)}), 500


@app.route("/api/matters/<file_no>/timeline", methods=["POST"])
def api_add_timeline(file_no):
    data = request.json or {}
    entry_date = data.get("date", datetime.now().strftime("%Y-%m-%d"))
    entry_text = data.get("text", "")
    if not entry_text:
        return jsonify({"error": "Text required"}), 400
    return jsonify({"success": add_timeline_entry(file_no, entry_date, entry_text)})


@app.route("/api/matters/<file_no>/next-action", methods=["POST"])
def api_update_next_action(file_no):
    return jsonify({"success": write_cell(file_no, "Next Action / Deadline", (request.json or {}).get("text", ""))})


@app.route("/api/matters/<file_no>/contact", methods=["POST"])
def api_update_contact(file_no):
    data = request.json or {}
    updates = {}
    for field, col in [("email", "Client Email"), ("phone", "Client Phone"), ("address", "Client Address"), ("other_parties", "Other Parties / Related Persons")]:
        if field in data:
            updates[col] = data[field]
    return jsonify({"success": write_cells(file_no, updates) if updates else True})


@app.route("/api/matters/<file_no>/limitation", methods=["POST"])
def api_update_limitation(file_no):
    data = request.json or {}
    discovery = data.get("discovery_date", "")
    statute = data.get("statute", "")
    deadline = data.get("deadline", "")
    if discovery and statute and not deadline:
        deadline = calc_limitation_deadline(discovery, statute) or ""
    write_cells(file_no, {
        "Discovery Date": discovery,
        "Limitation Statute": statute,
        "Limitation Deadline": deadline,
    })
    return jsonify({"success": True, "calculated_deadline": deadline})


@app.route("/api/matters/<file_no>/limitation/clear", methods=["POST"])
def api_clear_limitation(file_no):
    """Remove limitation tracking from a matter."""
    write_cells(file_no, {
        "Discovery Date": "",
        "Limitation Statute": "",
        "Limitation Deadline": "",
    })
    return jsonify({"success": True})


@app.route("/api/matters/<file_no>/court-deadlines", methods=["POST"])
def api_add_court_deadline(file_no):
    """Add a single court-ordered deadline."""
    data = request.json or {}
    dl_date = data.get("date", "")
    description = data.get("description", "")
    source = data.get("source", "")
    if not dl_date or not description:
        return jsonify({"error": "Date and description required"}), 400

    # Load existing deadlines
    matters = load_matters()
    current = []
    for m in matters:
        if str(m.get("File #")) == str(file_no):
            current = parse_json_field(m.get("Court Deadlines"))
            break

    current.append({"date": dl_date, "description": description, "source": source})
    current.sort(key=lambda x: x.get("date", ""))
    write_cell(file_no, "Court Deadlines", json.dumps(current))
    return jsonify({"success": True, "deadlines": current})


@app.route("/api/matters/<file_no>/court-deadlines/remove", methods=["POST"])
def api_remove_court_deadline(file_no):
    """Remove a court deadline by index."""
    data = request.json or {}
    idx = data.get("index", -1)

    matters = load_matters()
    current = []
    for m in matters:
        if str(m.get("File #")) == str(file_no):
            current = parse_json_field(m.get("Court Deadlines"))
            break

    if 0 <= idx < len(current):
        current.pop(idx)
        write_cell(file_no, "Court Deadlines", json.dumps(current) if current else "")
        return jsonify({"success": True, "deadlines": current})
    return jsonify({"error": "Invalid index"}), 400


@app.route("/api/conflict-check", methods=["POST"])
def api_conflict_check():
    data = request.json or {}
    query = data.get("query", "").strip().lower()
    if not query or len(query) < 2:
        return jsonify({"error": "Query too short"}), 400
    terms = [t.strip() for t in query.split(",") if t.strip()]
    matters = load_matters()
    hits = []
    for m in matters:
        d = matter_to_dict(m)
        searchable_fields = {
            "Client": d["client"],
            "Opposing Party": d["opposing"],
            "Description": d["description"],
            "Timeline": d["timeline_raw"],
            "Other Parties": d["other_parties"],
            "Client Email": d["client_email"],
            "Client Phone": d["client_phone"],
            "Client Address": d["client_address"],
        }
        matched_terms = []
        locations = set()
        for term in terms:
            for field_name, field_val in searchable_fields.items():
                if term in field_val.lower():
                    matched_terms.append(term)
                    locations.add(field_name)
        if matched_terms:
            hits.append({
                "file_no": d["file_no"],
                "client": d["client"],
                "opposing": d["opposing"],
                "other_parties": d["other_parties"],
                "status": d["status"],
                "matter_type": d["matter_type"],
                "description": d["description"][:150],
                "matched_terms": list(set(matched_terms)),
                "locations": list(locations),
            })
    return jsonify({"query": query, "terms": terms, "hits": hits, "total_searched": len(matters)})


@app.route("/api/matters/<file_no>/folder", methods=["POST"])
def api_update_folder(file_no):
    path = (request.json or {}).get("path", "")
    return jsonify({"success": write_cell(file_no, "Matter Folder", path)})


@app.route("/api/matters/<file_no>/folder/open", methods=["POST"])
def api_open_folder(file_no):
    """Open the matter folder in the system file manager."""
    matters = load_matters()
    subfolder = ""
    for m in matters:
        if str(m.get("File #")) == str(file_no):
            subfolder = str(m.get("Matter Folder", "") or "")
            break
    if not subfolder:
        return jsonify({"error": "No folder linked"}), 404
    full_path = os.path.realpath(os.path.join(MATTER_FOLDER_BASE, subfolder))
    # Prevent path traversal outside the base directory
    if not full_path.startswith(os.path.realpath(MATTER_FOLDER_BASE)):
        return jsonify({"error": "Invalid folder path"}), 400
    if not os.path.isdir(full_path):
        return jsonify({"error": f"Folder not found: {subfolder}"}), 404
    # macOS: open; Linux: xdg-open; Windows: explorer
    import platform
    opener = {"Darwin": "open", "Linux": "xdg-open"}.get(platform.system(), "explorer")
    subprocess.Popen([opener, full_path])
    return jsonify({"success": True})


def generate_next_file_no():
    """Generate the next sequential file number in YYYY-NNN format."""
    year = datetime.now().strftime("%Y")
    matters_list = load_matters()
    max_seq = 0
    for m in matters_list:
        fn = safe_str(m.get("File #"))
        if fn.startswith(year + "-"):
            try:
                seq = int(fn.split("-", 1)[1])
                if seq > max_seq:
                    max_seq = seq
            except (ValueError, IndexError):
                pass
    return f"{year}-{max_seq + 1:03d}"


def append_matter_row(data):
    """Append a new matter row to the Open Matters sheet."""
    file_no = generate_next_file_no()
    try:
        backup_spreadsheet()
        wb = load_workbook(XLSX_PATH)
    except Exception as e:
        app.logger.error(f"Failed to open spreadsheet for new matter: {e}")
        return None
    try:
        ws = wb["Open Matters"]
        headers = [cell.value for cell in ws[1]]
        new_row = ws.max_row + 1
        col_vals = {
            "File #": file_no,
            "Client Name": data.get("client_name", ""),
            "Matter Description": data.get("description", ""),
            "Status": "Open",
            "Date Opened": datetime.now().strftime("%Y-%m-%d"),
            "Opposing Party": data.get("opposing", ""),
            "Client Email": data.get("email", ""),
            "Client Phone": data.get("phone", ""),
        }
        for col_name, val in col_vals.items():
            if col_name in headers and val:
                ws.cell(row=new_row, column=headers.index(col_name) + 1).value = val
        wb.save(XLSX_PATH)
        wb.close()
        return file_no
    except Exception as e:
        app.logger.error(f"Failed to append new matter: {e}")
        wb.close()
        return None


def move_matter_between_sheets(file_no, from_sheet, to_sheet, updates):
    """Move a matter row from one sheet to another, applying updates."""
    try:
        backup_spreadsheet()
        wb = load_workbook(XLSX_PATH)
    except Exception as e:
        app.logger.error(f"Failed to open spreadsheet for move: {e}")
        return False
    try:
        if from_sheet not in wb.sheetnames or to_sheet not in wb.sheetnames:
            wb.close()
            return False
        src = wb[from_sheet]
        dst = wb[to_sheet]
        src_headers = [cell.value for cell in src[1]]
        dst_headers = [cell.value for cell in dst[1]]
        file_col = src_headers.index("File #") + 1 if "File #" in src_headers else None
        if not file_col:
            wb.close()
            return False
        # Find source row
        src_row = None
        for row in range(2, src.max_row + 1):
            if str(src.cell(row=row, column=file_col).value) == str(file_no):
                src_row = row
                break
        if not src_row:
            wb.close()
            return False
        # Read all values from source row
        row_data = {}
        for i, h in enumerate(src_headers):
            row_data[h] = src.cell(row=src_row, column=i + 1).value
        # Apply updates
        for k, v in updates.items():
            row_data[k] = v
        # Append to destination sheet
        dst_row = dst.max_row + 1
        for col_name, val in row_data.items():
            if col_name in dst_headers:
                dst.cell(row=dst_row, column=dst_headers.index(col_name) + 1).value = val
        # Delete source row
        src.delete_rows(src_row)
        wb.save(XLSX_PATH)
        wb.close()
        return True
    except Exception as e:
        app.logger.error(f"Failed to move matter {file_no}: {e}")
        wb.close()
        return False


@app.route("/api/matters/new", methods=["POST"])
def api_create_matter():
    data = request.json or {}
    client_name = data.get("client_name", "").strip()
    description = data.get("description", "").strip()
    if not client_name or not description:
        return jsonify({"error": "Client name and description are required"}), 400
    file_no = append_matter_row(data)
    if not file_no:
        return jsonify({"error": "Failed to create matter"}), 500
    return jsonify({"success": True, "file_no": file_no})


@app.route("/api/matters/<file_no>/close", methods=["POST"])
def api_close_matter(file_no):
    today_str = datetime.now().strftime("%Y-%m-%d")
    success = move_matter_between_sheets(file_no, "Open Matters", "Closed Matters", {
        "Status": "Closed",
        "Date Closed": today_str,
    })
    if not success:
        return jsonify({"error": "Failed to close matter"}), 500
    add_timeline_entry(file_no, today_str, "Matter closed")
    return jsonify({"success": True})


@app.route("/api/matters/<file_no>/reopen", methods=["POST"])
def api_reopen_matter(file_no):
    today_str = datetime.now().strftime("%Y-%m-%d")
    success = move_matter_between_sheets(file_no, "Closed Matters", "Open Matters", {
        "Status": "Open",
        "Date Closed": "",
    })
    if not success:
        return jsonify({"error": "Failed to reopen matter"}), 500
    add_timeline_entry(file_no, today_str, "Matter reopened")
    return jsonify({"success": True})


@app.route("/api/matters/<file_no>/folder/suggest")
def api_suggest_folders(file_no):
    """Suggest matter folders based on client name fuzzy matching."""
    matters_list = load_matters()
    client_name = ""
    for m in matters_list:
        if str(m.get("File #")) == str(file_no):
            client_name = safe_str(m.get("Client Name"))
            break
    if not client_name:
        return jsonify({"suggestions": []})
    client_parts = [p.strip().lower() for p in re.split(r'[\s,]+', client_name) if len(p.strip()) >= 2]
    if not client_parts:
        return jsonify({"suggestions": []})
    suggestions = []
    try:
        for dirname in os.listdir(MATTER_FOLDER_BASE):
            full = os.path.join(MATTER_FOLDER_BASE, dirname)
            if not os.path.isdir(full):
                continue
            dir_lower = dirname.lower()
            score = sum(1 for part in client_parts if part in dir_lower)
            if score > 0:
                suggestions.append({"folder_name": dirname, "score": score})
    except OSError:
        pass
    suggestions.sort(key=lambda x: -x["score"])
    return jsonify({"suggestions": suggestions[:5]})


@app.route("/api/matters/<file_no>/type", methods=["POST"])
def api_update_type(file_no):
    data = request.json or {}
    matter_type = data.get("type", "")
    return jsonify({"success": write_cell(file_no, "Matter Type", matter_type)})


@app.route("/api/statutes")
def api_statutes():
    return jsonify(LIMITATION_STATUTES)


if __name__ == "__main__":
    # Customise these for your firm
    FIRM_NAME = os.environ.get("FIRM_NAME", "My Law Firm")
    print(f"\n  {FIRM_NAME} — Matter Tracker")
    print(f"  Reading from: {XLSX_PATH}")
    print(f"  Backups saved to: {BACKUP_DIR}")
    print(f"  Open http://localhost:5001 in your browser\n")
    app.run(debug=os.environ.get("FLASK_DEBUG", "0") == "1", port=5001)
