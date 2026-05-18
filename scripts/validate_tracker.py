#!/usr/bin/env python3
"""Post-write integrity check for matter-tracker.xlsx.

Usage:
    python3 validate_tracker.py <tracker_path> [backup_path]

Exit codes: 0 = pass, 1 = fail (details on stderr).
When backup_path is given, also checks that row count didn't decrease.
"""

import sys
from pathlib import Path

EXPECTED_HEADERS = [
    "File #", "Client Name", "Matter Description", "Status",
    "Date Opened", "Date Closed", "Last Activity", "Opposing Party",
    "Next Action / Deadline", "Timeline", "Client ID Verified",
    "Conflict Check Done", "Client Email", "Client Phone",
    "Client Address", "Discovery Date", "Limitation Statute",
    "Limitation Deadline", "Court Deadlines", "Matter Folder",
    "Other Parties / Related Persons", "Matter Type",
]
VALID_STATUSES = {"Open", "Closed"}
MAX_COL = len(EXPECTED_HEADERS)  # 22 = V

errors = []


def err(msg):
    errors.append(msg)


def count_data_rows(ws):
    return max(0, ws.max_row - 1)


def validate_sheet(ws, sheet_name):
    headers = [ws.cell(row=1, column=c).value for c in range(1, MAX_COL + 1)]
    for i, (got, want) in enumerate(zip(headers, EXPECTED_HEADERS)):
        if got != want:
            err(f"[{sheet_name}] Column {chr(65+i)} header: expected '{want}', got '{got}'")

    for row in range(2, ws.max_row + 1):
        file_num = ws.cell(row=row, column=1).value
        status = ws.cell(row=row, column=4).value
        client = ws.cell(row=row, column=2).value

        all_none = all(ws.cell(row=row, column=c).value is None for c in range(1, MAX_COL + 1))
        if all_none:
            continue

        if not file_num:
            err(f"[{sheet_name}] Row {row}: blank File #  (Client: {client})")

        if status and status not in VALID_STATUSES:
            err(f"[{sheet_name}] Row {row}: invalid Status '{status}' (expected Open/Closed)")

        if sheet_name == "Open Matters" and status == "Closed":
            err(f"[{sheet_name}] Row {row}: Status is 'Closed' but row is on Open Matters (File #{file_num})")
        if sheet_name == "Closed Matters" and status == "Open":
            err(f"[{sheet_name}] Row {row}: Status is 'Open' but row is on Closed Matters (File #{file_num})")

    file_nums = []
    for row in range(2, ws.max_row + 1):
        v = ws.cell(row=row, column=1).value
        if v:
            file_nums.append(v)
    seen = set()
    for fn in file_nums:
        if fn in seen:
            err(f"[{sheet_name}] Duplicate File #: {fn}")
        seen.add(fn)


def main():
    if len(sys.argv) < 2:
        print(__doc__, file=sys.stderr)
        sys.exit(1)

    tracker_path = Path(sys.argv[1])
    backup_path = Path(sys.argv[2]) if len(sys.argv) > 2 else None

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("FAIL: openpyxl not installed", file=sys.stderr)
        sys.exit(1)

    try:
        wb = load_workbook(tracker_path, data_only=True)
    except Exception as e:
        print(f"FAIL: cannot open {tracker_path}: {e}", file=sys.stderr)
        sys.exit(1)

    if "Open Matters" not in wb.sheetnames:
        err("Missing sheet: 'Open Matters'")
    else:
        validate_sheet(wb["Open Matters"], "Open Matters")

    if "Closed Matters" in wb.sheetnames:
        validate_sheet(wb["Closed Matters"], "Closed Matters")

    if backup_path and backup_path.exists():
        try:
            bk = load_workbook(backup_path, data_only=True)
            for sheet_name in ["Open Matters", "Closed Matters"]:
                if sheet_name in bk.sheetnames and sheet_name in wb.sheetnames:
                    old_count = count_data_rows(bk[sheet_name])
                    new_count = count_data_rows(wb[sheet_name])
                    total_old = sum(count_data_rows(bk[s]) for s in ["Open Matters", "Closed Matters"] if s in bk.sheetnames)
                    total_new = sum(count_data_rows(wb[s]) for s in ["Open Matters", "Closed Matters"] if s in wb.sheetnames)
            if total_new < total_old:
                err(f"Total row count decreased: {total_old} → {total_new} (possible data loss)")
        except Exception as e:
            err(f"Could not compare against backup: {e}")

    wb.close()

    if errors:
        print(f"FAIL: {len(errors)} issue(s) found:", file=sys.stderr)
        for e in errors:
            print(f"  - {e}", file=sys.stderr)
        sys.exit(1)
    else:
        print("PASS: tracker integrity OK", file=sys.stderr)
        sys.exit(0)


if __name__ == "__main__":
    main()
